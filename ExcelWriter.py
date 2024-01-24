from openpyxl import Workbook

class ExcelWriter:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.create_cell(0, "Result")

    def set_cell_data(self, data):
        row_num = self.sheet.max_row + 1
        self.create_cell(0, data, row_num)

    def save(self):
        self.workbook.save(self.file_path)

    def create_cell(self, col_num, data, row_num=None):
        if row_num is None:
            row_num = self.sheet.max_row
        cell = self.sheet.cell(row=row_num, column=col_num + 1)
        cell.value = data

# Example usage:
excel_writer = ExcelWriter("output.xlsx")
excel_writer.set_cell_data("Example Data 1")
excel_writer.set_cell_data("Example Data 2")
excel_writer.save()
