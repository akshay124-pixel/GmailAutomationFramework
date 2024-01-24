from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager


class ExcelReader:
    def __init__(self, file_path):
        self.workbook = load_workbook(file_path)
        self.sheet = self.workbook.active

    def get_cell_data(self, row_num, col_num):
        cell = self.sheet.cell(row=row_num, column=col_num)
        return cell.value

    def get_row_count(self):
        return self.sheet.max_row

class ExcelWriter:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = load_workbook(file_path)
        self.sheet = self.workbook.active
        self.current_row = 1

    def set_cell_data(self, data):
        self.sheet.cell(row=self.current_row, column=1, value=data)
        self.current_row += 1

    def save(self):
        self.workbook.save(self.file_path)

def login_to_gmail(driver, username, password, excel_writer):
    # Locate and enter username
    username_input = driver.find_element(By.ID, "identifierId")
    username_input.send_keys(username)
    username_input.send_keys(Keys.RETURN)

    # Wait for the password input field
    password_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.NAME, "password"))
    )
    password_input.send_keys(password)
    password_input.send_keys(Keys.RETURN)

    # Check if login is successful
    login_successful = "inbox" in driver.title.lower()

    # Write result to Excel
    excel_writer.set_cell_data("Success" if login_successful else "Failure")

# Set the path to ChromeDriver executable
chrome_driver_path = "path/to/chromedriver"
# Launch Chrome browser
driver = webdriver.Chrome(ChromeDriverManager().install())
try:
    # Navigate to Gmail login page
    driver.get("https://mail.google.com/")

    # Your login credentials Excel file path
    excel_file_path = "path/to/login_credentials.xlsx"

    # Create an instance of ExcelReader to read data from the Excel file
    excel_reader = ExcelReader(excel_file_path)

    # Your Excel file path for result
    result_excel_file_path = "path/to/result.xlsx"

    # Create an instance of ExcelWriter to write results to the Excel file
    excel_writer = ExcelWriter(result_excel_file_path)

    # Read and process each row of login credentials
    for i in range(1, excel_reader.get_row_count() + 1):
        username = excel_reader.get_cell_data(i, 1)
        password = excel_reader.get_cell_data(i, 2)

        # Execute login process
        login_to_gmail(driver, username, password, excel_writer)

finally:
    # Close the browser after all operations
    driver.quit()
    # Save the result Excel file
    excel_writer.save()
