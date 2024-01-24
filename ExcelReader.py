from openpyxl import load_workbook

class ExcelReader:
    def __init__(self, file_path):
        try:
            self.workbook = load_workbook(file_path)
            self.sheet = self.workbook.active
        except Exception as e:
            print(f"Error: {e}")

    def get_cell_data(self, row_num, col_num):
        cell = self.sheet.cell(row=row_num, column=col_num)
        return cell.value

    def get_row_count(self):
        return self.sheet.max_row

# Example usage:
excel_reader = ExcelReader("your_file.xlsx")
print(f"Row count: {excel_reader.get_row_count()}")

for row_number in range(1, excel_reader.get_row_count() + 1):
    for col_number in range(1, excel_reader.sheet.max_column + 1):
        cell_data = excel_reader.get_cell_data(row_number, col_number)
        print(f"Row {row_number}, Column {col_number}: {cell_data}")
